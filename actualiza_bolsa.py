
import sys
import os
import math
import datetime as dt
from datetime import datetime, timedelta
from openpyxl import load_workbook

try:
    import pandas as pd
except ImportError:
    print("[ERROR] Falta pandas. Instala con: pip install pandas openpyxl yfinance")
    sys.exit(1)

try:
    import yfinance as yf
except ImportError:
    print("[ERROR] Falta yfinance. Instala con: pip install yfinance")
    sys.exit(1)

# -----------------------------
# Configuración
# -----------------------------
TICKERS = {
    "Repsol": "REP.MC",
    "Wolters Kluwer": "WKL.AS",
    "Apple": "AAPL",
    "Telefónica": "TEF", 
    "Banco Santander": "SAN",
    "Banco Sabadell": "SAB",
    "BBVA": "BBVA",
    "OHLA": "OHLA.MC",
    "Sacyr": "SYV",
    "AENA": "AENA",
    "Acciona": "ANA",
    "Ferrovial": "FER",
    "Atos": "ATO.PA",
    "Alten": "ATE.PA",
    "Sopra Steria": "SOP.PA",
    "Indra": "IDR.MC",
    "Amadeus": "AMS",
    "Grifols": "GRF",
    "Inditex (Zara)": "ITX.MC",
    "YPF": "YPF",
    "Iberdrola": "IBE",
}

SALIDA_XLSX = "bolsa_resumen_auto.xlsx"
HOJA = "Resumen"

# Columnas solicitadas
COLUMNAS = [
    "Nombre", "Ticker", "Mercado", "Fecha", "Valor actual", "Δ ayer",
    "Δ semana", "Δ 3 meses", "Mínimo mes", "Mínimo año", "Máximo mes", "Máximo año",
    "Máximo absoluto", "Última noticia relevante"
]

# -----------------------------
# Utilidades de fechas
# -----------------------------

def last_close_before(df, target_date):
    """Devuelve el último cierre disponible en o antes de target_date (fecha calendario)."""
    if df.empty:
        return None

    # Asegurar índice tipo datetime y ordenado
    s = df.sort_index()
    s = s.loc[s.index <= pd.to_datetime(target_date)]

    if s.empty:
        return None

    return float(s["Close"].iloc[-1].squeeze())


def pct(a, b):
    """Variación porcentual (a vs b). Devuelve string con % y 2 decimales."""
    if a is None or b is None or b == 0 or (isinstance(b, float) and (math.isnan(b) or b == 0)):
        return ""

    return f"{(a/b - 1)*100:.2f}%"


# -----------------------------
# Descarga y cálculo por ticker
# -----------------------------
import pandas as pd

resumen_rows = []

print(f"Mover datos antiguos para hacer hueco a los nuevos datos.")
wb = load_workbook(SALIDA_XLSX)
ws = wb["Resumen"]
ultima_fila = ws.max_row
print(f"Obtener última línea:", ultima_fila)
numTickers = len(TICKERS)
ws.insert_rows(2, amount=numTickers)
wb.save(SALIDA_XLSX)

# Obtengo la fecha de hoy
hoy_str = datetime.today().strftime("%d/%m/%Y")

for nombre, ticker in TICKERS.items():
    print(f"Procesando {nombre} ({ticker})...")
    try:
        # Serie 400 días para cálculos 1w, 1m, 3m, 1y
        hist = yf.download(ticker, period="400d", interval="1d", auto_adjust=False, progress=False)
        # Serie completa para máximo histórico
        hist_all = yf.download(ticker, period="max", interval="1d", auto_adjust=False, progress=False)
    except Exception as e:
        print(f"  [WARN] No se pudo descargar {ticker}: {e}")
        hist = pd.DataFrame()
        hist_all = pd.DataFrame()

    # Mercado (sufijo simplificado por exchange)
    mercado = ""
    if ticker.endswith(".MC"):
        mercado = "BME"
    elif ticker.endswith(".PA"):
        mercado = "EPA"
    elif ticker.endswith(".AS"):
        mercado = "AMS"
    elif ticker.upper() == "AAPL":
        mercado = "NASDAQ"
    elif ticker.upper() == "YPF":
        mercado = "NYSE"

    # Valor actual y variación diaria
    valor_actual = ""
    var_dia = ""
    if not hist.empty:
        last_close = float(hist["Close"].iloc[-1].squeeze())
        valor_actual = last_close
        if len(hist) >= 2:
            prev_close = float(hist["Close"].iloc[-2].squeeze())
            var_dia = pct(last_close, prev_close)

    # Fechas objetivo
    hoy = pd.to_datetime(dt.date.today())
    hace_7 = hoy - pd.Timedelta(days=7)
    hace_90 = hoy - pd.Timedelta(days=90)
    hace_30 = hoy - pd.Timedelta(days=30)
    hace_365 = hoy - pd.Timedelta(days=365)

    # Variaciones 1 semana y 3 meses
    var_1w = ""
    var_3m = ""
    if not hist.empty:
        c_hoy = last_close_before(hist, hoy)
        c_1w = last_close_before(hist, hace_7)
        c_3m = last_close_before(hist, hace_90)
        var_1w = pct(c_hoy, c_1w)
        var_3m = pct(c_hoy, c_3m)

    # Mín/Máx 30 días y 365 días
    min_30 = ""; max_30 = ""; min_365 = ""; max_365 = ""; max_all = ""
    if not hist.empty:
        rec_30 = hist.loc[hist.index >= hace_30]
        if not rec_30.empty:
            min_30 = float(rec_30["Low"].min().squeeze())
            max_30 = float(rec_30["High"].max().squeeze())
        rec_365 = hist.loc[hist.index >= hace_365]
        if not rec_365.empty:
            min_365 = float(rec_365["Low"].min().squeeze())
            max_365 = float(rec_365["High"].max().squeeze())
    if not hist_all.empty:
        max_all = float(hist_all["High"].max().squeeze())

    # Última noticia
    noticia = ""
    try:
        t = yf.Ticker(ticker)
        news = getattr(t, 'news', None)
        if news:
            first = news[0]
            tit = first.get('title', '')
            pub = first.get('publisher', '')
            ts = first.get('providerPublishTime') or first.get('publishedAt')
            if ts:
                fecha = dt.datetime.utcfromtimestamp(int(ts)).strftime('%Y-%m-%d') if isinstance(ts, (int, float, str)) else ''
            else:
                fecha = ''
            noticia = f"{tit} ({pub}, {fecha})"
    except Exception:
        pass

    # Formateo de números con símbolo monetario según mercado (aprox)
    def fmt(v):
        if v == "" or v is None:
            return ""
        if mercado in ("BME", "EPA", "AMS"):
            return f"€{v:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")
        else:
            return f"${v:,.2f}"

    fila = {
        "Nombre": nombre,
        "Ticker": ticker,
        "Mercado": mercado,
        "Fecha": hoy_str,
        "Valor actual": fmt(valor_actual) if valor_actual != "" else "",
        "Δ ayer": var_dia,
        "Δ semana": var_1w,
        "Δ 3 meses": var_3m,
        "Mínimo mes": fmt(min_30),
        "Mínimo año": fmt(min_365),
        "Máximo mes": fmt(max_30),
        "Máximo año": fmt(max_365),
        "Máximo absoluto": fmt(max_all),
        "Última noticia relevante": noticia,
    }
    resumen_rows.append(fila)

# Construcción del DataFrame final
resumen_df = pd.DataFrame(resumen_rows, columns=COLUMNAS)

# Guardar Excel (una sola hoja "Resumen")
with pd.ExcelWriter(SALIDA_XLSX, engine="openpyxl", mode="a",
    if_sheet_exists="overlay"  # o "overlay"/"new" según lo que quieras
    ) as writer:
    resumen_df.to_excel(writer, sheet_name=HOJA, index=False)
    
    # Ajustar el ancho de las celdas.
    hoja = writer.sheets[HOJA]
    for col_idx, col_name in enumerate(resumen_df.columns, start=1):
        max_len = max(
            [len(str(col_name))] +
            [len(str(v)) for v in resumen_df[col_name].values]
        )
        hoja.column_dimensions[hoja.cell(row=1, column=col_idx).column_letter].width = max_len + 2
print(f"[OK] Archivo actualizado: {SALIDA_XLSX}")
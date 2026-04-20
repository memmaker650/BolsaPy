
import sys
import os
import math
import time
import datetime as dt
import concurrent.futures
from datetime import datetime, timedelta
from openpyxl import load_workbook

import pandas as pd

import sqlite3
from pathlib import Path


class ActualizaBolsa:
    conn = None
    resumen_df = None
    db_path = None

    try:
        import pandas as pd
    except ImportError:
        print("[ERROR] Falta pandas. Instala con: pip install pandas openpyxl yfinance")
        sys.exit(1)

    try:
        import yfinance as yf
    except ImportError:
        print("[ERROR] Falta yfinance. Instala con: pip install yfinance")
        sys.exit(1)

    # -----------------------------
    # TICKERS
    # -----------------------------
    TICKERS = {
        "Repsol": "REP.MC",
        "Wolters Kluwer": "WKL.AS",
        "Apple": "AAPL",
        "Telefónica": "TEF.MC", 
        "Amadeus": "AMS.MC",
        "Banco Santander": "SAN",
        "Banco Sabadell": "SAB.MC",
        "BBVA": "BBVA",
        "OHLA": "OHLA.MC",
        "Sacyr": "SCYR.MC",
        "AENA": "AENA.MC",
        "Acciona": "ANA.MC",
        "Ferrovial": "FER",
        "Atos": "ATO.PA",
        "Alten": "ATE.PA",
        "Sopra Steria": "SOP.PA",
        "Indra": "IDR.MC",
        "Amadeus": "AMS",
        "Grifols": "GRF",
        "Inditex (Zara)": "ITX.MC",
        "Repsol": "REP.MC",
        "YPF": "YPF",
        "Endesa": "ELE.MC",
        "Iberdrola": "IBE.MC",
        "NIKE": "NKE",
        "ADIDAS": "ADS.DE",
        "ASICS": "7936.T",
    }

    # -----------------------------
    # OPERACIONES CON BDD
    # -----------------------------
    def conectarBDD(self):
        self.conn = sqlite3.connect(self.db_path)
        cursor = self.conn.cursor()

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS valores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT,
            TICKERS TEXT,
            Mercado TEXT, 
            Fecha Date, 
            Valor_actual NUMERIC, 
            Delta_ayer NUMERIC,
            Delta_semana NUMERIC, 
            Delta_3meses NUMERIC, 
            MinimoMes NUMERIC, 
            Minimo_Agno NUMERIC, 
            Maximo_Mes NUMERIC, 
            Maximo_Agno NUMERIC,
            Maximo_Absoluto NUMERIC, 
            Ultima_Noticia_Relevante NUMERIC
        )""")

        self.conn.commit()

    def escribirBDD(self, x1, x2, x3, x4, x5, x6, x7, x8, x9, x10, x11, x12, x13, x14):
        cursor = self.conn.cursor()

        cursor.execute("""
            INSERT INTO valores (
            nombre,
            TICKERS,
            Mercado, 
            Fecha, 
            Valor_actual, 
            Delta_ayer,
            Delta_semana, 
            Delta_3meses, 
            MinimoMes, 
            Minimo_Agno, 
            Maximo_Mes, 
            Maximo_Agno,
            Maximo_Absoluto , 
            Ultima_Noticia_Relevante
        ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (x1, x2, x3, x4, x5, x6, x7, x8, x9, x10, x11, x12, x13, x14))


    def cerrarBDD(self):
        self.conn.close()


    SALIDA_XLSX = "bolsa_resumen_auto.xlsx"
    HOJA = "Resumen"

    # Columnas solicitadas
    COLUMNAS = [
        "Nombre", "Ticker", "Mercado", "Fecha", "Valor actual", "Δ ayer",
        "Δ semana", "Δ 3 meses", "Mínimo mes", "Mínimo año", "Máximo mes", "Máximo año",
        "Máximo absoluto", "Última noticia relevante"
    ]

    # Utilidades de fechas
    # -----------------------------
    # -----------------------------
    def last_close_before(self, df, target_date):
        """Devuelve el último cierre disponible en o antes de target_date (fecha calendario)."""
        if df.empty:
            return None

        # Asegurar índice tipo datetime y ordenado
        s = df.sort_index()
        idx = s.index
        target_ts = pd.to_datetime(target_date)
        # Evita comparar tz-aware vs tz-naive (p. ej. Europe/Madrid vs Timestamp naive).
        if getattr(idx, "tz", None) is not None and target_ts.tzinfo is None:
            target_ts = target_ts.tz_localize(idx.tz)
        elif getattr(idx, "tz", None) is None and target_ts.tzinfo is not None:
            target_ts = target_ts.tz_localize(None)

        s = s.loc[idx <= target_ts]

        if s.empty:
            return None

        return float(s["Close"].iloc[-1].squeeze())


    def pct(self, a, b):
        """Variación porcentual (a vs b). Devuelve string con % y 2 decimales."""
        if a is None or b is None or b == 0 or (isinstance(b, float) and (math.isnan(b) or b == 0)):
            return ""

        return f"{(a/b - 1)*100:.2f}%"

    def _normalize_history_index(self, df):
        if isinstance(df, pd.DataFrame) and not df.empty and getattr(df.index, "tz", None) is not None:
            df = df.copy()
            df.index = df.index.tz_localize(None)
        return df

    def _get_history_safe(self, ticker, period="max"):
        """
        Descarga robusta con reintento y fallback.
        yfinance a veces devuelve TypeError('NoneType'...) de forma intermitente.
        """
        last_error = None
        for attempt in range(2):
            try:
                tk = self.yf.Ticker(ticker)
                hist = tk.history(period=period, interval="1d", auto_adjust=False, timeout=12)
                return tk, self._normalize_history_index(hist)
            except Exception as e:
                last_error = e
                time.sleep(0.5 * (attempt + 1))

        try:
            # Fallback alternativo cuando history() falla internamente.
            hist = self.yf.download(
                ticker, period=period, interval="1d", auto_adjust=False, progress=False, threads=False
            )
            return self.yf.Ticker(ticker), self._normalize_history_index(hist)
        except Exception as e:
            last_error = e
            print(f"  [WARN] No se pudo descargar {ticker}: {last_error}")
            return self.yf.Ticker(ticker), pd.DataFrame()


    # -----------------------------
    # Descarga y cálculo por ticker
    # -----------------------------
    def descargaYCalculoTickers(self):
        resumen_rows = []

        print(f"Mover datos antiguos para hacer hueco a los nuevos datos.")
        wb = load_workbook(self.SALIDA_XLSX)
        ws = wb["Resumen"]
        ultima_fila = ws.max_row
        print(f"Obtener última línea:", ultima_fila)
        numTickers = len(self.TICKERS)
        ws.insert_rows(2, amount=numTickers)
        wb.save(self.SALIDA_XLSX)

        # Obtengo la fecha de hoy
        hoy_str = datetime.today().strftime("%d/%m/%Y")

        for nombre, ticker in self.TICKERS.items():
            print(f"Procesando {nombre} ({ticker})...")
            tk, hist_all = self._get_history_safe(ticker, period="max")
            hist = hist_all.tail(400) if not hist_all.empty else pd.DataFrame()

            # Mercado (sufijo simplificado por exchange)
            mercado = ""
            if ticker.endswith(".MC"):
                mercado = "BME"
            elif ticker.endswith(".PA"):
                mercado = "EPA"
            elif ticker.endswith(".AS"):
                mercado = "AMS"
            elif ticker.upper() == "AAPL":
                mercado = "NASDAQ"
            elif ticker.upper() == "YPF":
                mercado = "NYSE"

            # Valor actual y variación diaria
            valor_actual = ""
            var_dia = ""
            if not hist.empty:
                last_close = float(hist["Close"].iloc[-1].squeeze())
                valor_actual = last_close
                if len(hist) >= 2:
                    prev_close = float(hist["Close"].iloc[-2].squeeze())
                    var_dia = self.pct(last_close, prev_close)

            # Fechas objetivo
            hoy = pd.to_datetime(dt.date.today())
            hace_7 = hoy - pd.Timedelta(days=7)
            hace_90 = hoy - pd.Timedelta(days=90)
            hace_30 = hoy - pd.Timedelta(days=30)
            hace_365 = hoy - pd.Timedelta(days=365)

            # Variaciones 1 semana y 3 meses
            var_1w = ""
            var_3m = ""
            if not hist.empty:
                c_hoy = self.last_close_before(hist, hoy)
                c_1w = self.last_close_before(hist, hace_7)
                c_3m = self.last_close_before(hist, hace_90)
                var_1w = self.pct(c_hoy, c_1w)
                var_3m = self.pct(c_hoy, c_3m)

            # Mín/Máx 30 días y 365 días
            min_30 = ""; max_30 = ""; min_365 = ""; max_365 = ""; max_all = ""
            if not hist.empty:
                rec_30 = hist.loc[hist.index >= hace_30]
                if not rec_30.empty:
                    min_30 = float(rec_30["Low"].min().squeeze())
                    max_30 = float(rec_30["High"].max().squeeze())
                rec_365 = hist.loc[hist.index >= hace_365]
                if not rec_365.empty:
                    min_365 = float(rec_365["Low"].min().squeeze())
                    max_365 = float(rec_365["High"].max().squeeze())
            if not hist_all.empty:
                max_all = float(hist_all["High"].max().squeeze())

            # Última noticia
            noticia = ""
            try:
                def _fetch_news():
                    return getattr(tk, "news", None)

                # Evita que una llamada lenta de noticias bloquee todo el proceso.
                with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
                    future = executor.submit(_fetch_news)
                    news = future.result(timeout=8)

                if news:
                    first = news[0]
                    tit = first.get('title', '')
                    pub = first.get('publisher', '')
                    ts = first.get('providerPublishTime') or first.get('publishedAt')
                    if ts:
                        fecha = dt.datetime.utcfromtimestamp(int(ts)).strftime('%Y-%m-%d') if isinstance(ts, (int, float, str)) else ''
                    else:
                        fecha = ''
                    noticia = f"{tit} ({pub}, {fecha})"
            except Exception:
                pass

            # Formateo de números con símbolo monetario según mercado (aprox)
            def fmt(v):
                if v == "" or v is None:
                    return ""
                if mercado in ("BME", "EPA", "AMS"):
                    return f"€{v:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")
                else:
                    return f"${v:,.2f}"

            fila = {
                "Nombre": nombre,
                "Ticker": ticker,
                "Mercado": mercado,
                "Fecha": hoy_str,
                "Valor actual": fmt(valor_actual) if valor_actual != "" else "",
                "Δ ayer": var_dia,
                "Δ semana": var_1w,
                "Δ 3 meses": var_3m,
                "Mínimo mes": fmt(min_30),
                "Mínimo año": fmt(min_365),
                "Máximo mes": fmt(max_30),
                "Máximo año": fmt(max_365),
                "Máximo absoluto": fmt(max_all),
                "Última noticia relevante": noticia,
            }
            resumen_rows.append(fila)

            self.escribirBDD(nombre, ticker, mercado, hoy_str, fmt(valor_actual) if valor_actual != "" else "", var_dia, var_1w, var_3m, fmt(min_30), fmt(min_365), fmt(max_30), fmt(max_365), fmt(max_all), noticia)

        # Construcción del DataFrame final
        self.resumen_df = pd.DataFrame(resumen_rows, columns=self.COLUMNAS)
        self.conn.commit()

    def guardadoEnExcel(self):
        # Guardar Excel (una sola hoja "Resumen")
        with pd.ExcelWriter(self.SALIDA_XLSX, engine="openpyxl", mode="a",
            if_sheet_exists="overlay"  # o "overlay"/"new" según lo que quieras
            ) as writer:
            self.resumen_df.to_excel(writer, sheet_name=self.HOJA, index=False)

            # Ajustar el ancho de las celdas.
            hoja = writer.sheets[self.HOJA]
            for col_idx, col_name in enumerate(self.resumen_df.columns, start=1):
                max_len = max(
                    [len(str(col_name))] +
                    [len(str(v)) for v in self.resumen_df[col_name].values]
                )
                hoja.column_dimensions[hoja.cell(row=1, column=col_idx).column_letter].width = max_len + 2
        print(f"[OK] Archivo actualizado: {self.SALIDA_XLSX}")

    def lanzarAcciones(self, widget=None):
        self.conectarBDD()
        try:
            self.descargaYCalculoTickers()
            self.guardadoEnExcel()
        finally:
            self.cerrarBDD()

def main():
    app = ActualizaBolsa()
    app.lanzarAcciones()
    
if __name__ == "__main__":
    main()